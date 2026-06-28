import io

from openpyxl import Workbook
from openpyxl.styles import Font


def facture_xlsx(facture):
    wb = Workbook()
    ws = wb.active
    ws.title = "Facture"
    gras = Font(bold=True)

    ws.append(["Facture de rétrocession", facture.numero or ""])
    ws.append(["Émettrice", facture.emettrice or ""])
    ws.append(["Destinataire", facture.destinataire or ""])
    ws.append(["Date", facture.date_vente or ""])
    if getattr(facture, "n_rouge", 0):
        ws.append(["FACTURE PARTIELLE",
                   f"{facture.n_rouge} ligne(s) non rapprochée(s) exclue(s) du total"])
    ws.append([])

    ws.append(["BL", "Date BL", "Désignation", "Code", "Qté", "PA brut", "Remise %",
               "PA net", "TVA", "Montant HT"])
    for cell in ws[ws.max_row]:
        cell.font = gras
    for g in facture.groupes:
        for l in g.lignes:
            ws.append([g.bl_numero, g.bl_date, l.designation, l.code, l.qte, l.prix_brut,
                       l.remise_pct, l.prix_net, l.tva, l.montant_ht])

    ws.append([])
    ws.append(["Ventilation TVA", "Taux", "Base HT", "Montant TVA"])
    for v in facture.ventilation:
        ws.append(["", v.taux, v.base_ht, v.montant_tva])

    ws.append([])
    for libelle, val in (("Total HT", facture.total_ht), ("Total TVA", facture.total_tva),
                         ("Total TTC", facture.total_ttc)):
        ws.append([libelle, val])
        ws[ws.max_row][0].font = gras

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
